
import sqlite3, os, sys, re

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH  = os.path.join(BASE_DIR, "smartsched.db")

DAYS      = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]
DAY_START = 8
DAY_END   = 20
TIMES     = list(range(DAY_START, DAY_END))

# ── Colour palette ────────────────────────────────────────────────────────────
NAVY    = "0F172A"
SLATE   = "1E293B"
SLATE2  = "334155"
STEEL   = "475569"
SKY_50  = "F0F9FF"
SKY_200 = "BAE6FD"
SKY_800 = "075985"
LAB_BG  = "DBEAFE"
LAB_FG  = "1E40AF"
LEC_BG  = "DCFCE7"
LEC_FG  = "166534"
AY_BG   = "EFF6FF"
AY_FG   = "0369A1"
TIME_BG = "F0F9FF"
TIME_FG = "075985"
WHITE   = "FFFFFF"
SEP_BG  = "E2E8F0"
BDR     = "BAE6FD"
BDR_DK  = "334155"
SEC_BG  = "4C1D95"   # deep purple
SEC_SUB = "6D28D9"   # medium purple

# ── Strip emoji ───────────────────────────────────────────────────────────────
EMOJI_RE = re.compile(
    "["
    u"\U0001F600-\U0001F64F"
    u"\U0001F300-\U0001F5FF"
    u"\U0001F680-\U0001F6FF"
    u"\U0001F1E0-\U0001F1FF"
    u"\U00002702-\U000027B0"
    u"\U000024C2-\U0001F251"
    "]+", flags=re.UNICODE
)
def strip_emoji(text): return EMOJI_RE.sub("", str(text)).strip()

# ── Style helpers ─────────────────────────────────────────────────────────────
def sd(s="thin", c=BDR):
    return Side(style=s, color=c)

def box(s="thin", c=BDR):
    x = sd(s, c)
    return Border(left=x, right=x, top=x, bottom=x)

def fill(c):
    return PatternFill("solid", fgColor=c)

def fnt(bold=False, sz=11, color="000000", italic=False):
    return Font(bold=bold, size=sz, color=color, italic=italic,
                name="Liberation Sans")

def aln(h="center", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap,
                     shrink_to_fit=False)

def fmt_h(h):
    if h == 0:  return "12:00 AM"
    if h == 12: return "12:00 PM"
    if h < 12:  return f"{h}:00 AM"
    return f"{h-12}:00 PM"

def fmt_r(s, e):
    return f"{fmt_h(s)} - {fmt_h(e)}"

def W(ws, row, col, val="", bold=False, sz=11, color="000000",
      bg=None, h="center", v="center", wrap=True, italic=False,
      bd=None, er=None, ec=None):
    if (ec and ec > col) or (er and er > row):
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=er or row, end_column=ec or col)
    c = ws.cell(row=row, column=col,
                value=strip_emoji(val) if isinstance(val, str) else val)
    c.font         = fnt(bold=bold, sz=sz, color=color, italic=italic)
    c.alignment    = aln(h=h, v=v, wrap=wrap)
    c.number_format = "@"
    if bg:
        c.fill = fill(bg)
    if bd:
        c.border = bd
    else:
        c.border = box("thin", "D1D5DB")
    return c

def blank_row(ws, row, c1, c2, bg=WHITE, h=6):
    ws.row_dimensions[row].height = h
    for c in range(c1, c2 + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill   = fill(bg)
        cell.border = box("thin", bg)

# ── Database ──────────────────────────────────────────────────────────────────
def get_data():
    conn = sqlite3.connect(DB_PATH)
    cur  = conn.cursor()
    cur.execute("""
        SELECT ss.section, ss.subject, ss.instructor, ss.day,
               ss.start_time, ss.end_time, ss.room, ss.room_type,
               COALESCE(ay.year,''), COALESCE(ay.semester,'')
        FROM student_schedules ss
        LEFT JOIN academic_years ay ON ss.academic_year_id = ay.id
        ORDER BY ss.section, ss.day, ss.start_time
    """)
    rows = cur.fetchall()
    conn.close()
    return [{"section": r[0], "subject": r[1], "instructor": r[2], "day": r[3],
             "start": r[4], "end": r[5], "room": r[6],
             "room_type": r[7], "ay": r[8], "sem": r[9]} for r in rows]


# ── Excel generator ───────────────────────────────────────────────────────────
def generate_excel(records):
    wb = Workbook()
    ws = wb.active
    ws.title = "Student Schedule"
    ws.sheet_view.showGridLines = False

    sections = list(dict.fromkeys(r["section"] for r in records))
    ay  = next((r["ay"]  for r in records if r["ay"]),  "")
    sem = next((r["sem"] for r in records if r["sem"]), "")

    TC = 1
    DC = {d: TC + 1 + i for i, d in enumerate(DAYS)}
    T  = TC + len(DAYS)

    ws.column_dimensions[get_column_letter(TC)].width = 18
    for d in DAYS:
        ws.column_dimensions[get_column_letter(DC[d])].width = 26

    ws.page_setup.orientation    = "landscape"
    ws.page_setup.paperSize      = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage      = False
    ws.page_setup.scale          = 75
    ws.page_margins = PageMargins(
        left=0.5, right=0.5, top=0.6, bottom=0.6, header=0.3, footer=0.3)
    ws.print_options.horizontalCentered = True

    row = 1

    # ── School header ─────────────────────────────────────────────
    blank_row(ws, row, TC, T, WHITE, 5); row += 1

    ws.row_dimensions[row].height = 48
    W(ws, row, TC, "PASSI CITY COLLEGE",
      bold=True, sz=28, color=NAVY, bg=WHITE,
      h="center", v="center", wrap=False, ec=T,
      bd=box("thin", WHITE))
    row += 1

    ws.row_dimensions[row].height = 24
    W(ws, row, TC, "Information and Communication Technology",
      bold=True, sz=14, color=SKY_800, bg=WHITE,
      h="center", v="center", wrap=False, ec=T,
      bd=box("thin", WHITE))
    row += 1

    ws.row_dimensions[row].height = 18
    W(ws, row, TC, "Passi City, Iloilo, Philippines",
      bold=False, sz=11, color=STEEL, bg=WHITE, italic=True,
      h="center", v="center", wrap=False, ec=T,
      bd=box("thin", WHITE))
    row += 1

    if ay or sem:
        ws.row_dimensions[row].height = 24
        ay_text = f"Academic Year {ay}   |   {sem}" if ay else sem
        for c in range(TC, T + 1):
            ws.cell(row=row, column=c).fill = fill(AY_BG)
        W(ws, row, TC, ay_text,
          bold=True, sz=12, color=AY_FG, bg=AY_BG,
          h="center", v="center", wrap=False, ec=T,
          bd=box("medium", SKY_200))
        row += 1

    thick = Side(style="medium", color=NAVY)
    for c in range(TC, T + 1):
        cl = ws.cell(row=row - 1, column=c)
        cl.border = Border(
            left=cl.border.left, right=cl.border.right,
            top=cl.border.top, bottom=thick)

    blank_row(ws, row, TC, T, WHITE, 8); row += 1

    ws.row_dimensions[row].height = 32
    W(ws, row, TC, "STUDENT CLASS SCHEDULE",
      bold=True, sz=18, color=NAVY, bg=WHITE,
      h="center", v="center", wrap=False, ec=T,
      bd=box("thin", WHITE))
    row += 1

    ws.row_dimensions[row].height = 3
    for c in range(TC, T + 1):
        cl = ws.cell(row=row, column=c)
        cl.fill   = fill(WHITE)
        cl.border = Border(bottom=Side(style="medium", color=NAVY))
    row += 1

    blank_row(ws, row, TC, T, WHITE, 10); row += 1

    header_start_row = row

    # ── Per-section tables ────────────────────────────────────────
    for sec in sections:
        cls = [r for r in records if r["section"] == sec]

        total_h = sum(r["end"] - r["start"] for r in cls)
        lab_h   = sum(r["end"] - r["start"] for r in cls if r["room_type"] == "Laboratory")
        lec_h   = sum(r["end"] - r["start"] for r in cls if r["room_type"] == "Lecture")
        lab_n   = sum(1 for r in cls if r["room_type"] == "Laboratory")
        lec_n   = sum(1 for r in cls if r["room_type"] == "Lecture")

        # Section name bar
        ws.row_dimensions[row].height = 36
        W(ws, row, TC, f"  {sec.upper()}",
          bold=True, sz=15, color=WHITE, bg=SEC_BG,
          h="left", v="center", wrap=False, ec=T,
          bd=box("medium", "3B0764"))
        row += 1

        # Hours summary bar
        ws.row_dimensions[row].height = 20
        summary = (
            f"  Total: {total_h} hr{'s' if total_h != 1 else ''}"
            f"   |   Lecture: {lec_h} hr{'s' if lec_h != 1 else ''} ({lec_n} slot{'s' if lec_n != 1 else ''})"
            f"   |   Laboratory: {lab_h} hr{'s' if lab_h != 1 else ''} ({lab_n} slot{'s' if lab_n != 1 else ''})"
        )
        W(ws, row, TC, summary,
          bold=False, sz=10, color="DDD6FE", bg=SEC_SUB,
          h="left", v="center", wrap=False, ec=T,
          bd=box("thin", SEC_BG))
        row += 1

        # Day header row
        ws.row_dimensions[row].height = 24
        W(ws, row, TC, "TIME",
          bold=True, sz=11, color=WHITE, bg=SLATE,
          h="center", v="center", wrap=False,
          bd=box("thin", BDR_DK))
        for d in DAYS:
            W(ws, row, DC[d], d,
              bold=True, sz=11, color=WHITE, bg=SLATE2,
              h="center", v="center", wrap=False,
              bd=box("thin", BDR_DK))
        row += 1

        # Time slot rows
        for hour in TIMES:
            ws.row_dimensions[row].height = 52   # taller for 4 lines

            tc = ws.cell(row=row, column=TC, value=fmt_r(hour, hour + 1))
            tc.font         = fnt(bold=True, sz=10, color=TIME_FG)
            tc.fill         = fill(TIME_BG)
            tc.alignment    = aln(h="center", v="center", wrap=False)
            tc.border       = box("thin", BDR)
            tc.number_format = "@"

            for d in DAYS:
                col   = DC[d]
                match = [r for r in cls
                         if r["day"] == d and r["start"] <= hour < r["end"]]
                dc = ws.cell(row=row, column=col)
                dc.border       = box("thin", BDR)
                dc.alignment    = aln(h="center", v="center", wrap=True)
                dc.number_format = "@"
                if match:
                    m      = match[0]
                    is_lab = m["room_type"] == "Laboratory"
                    inst_line = f"\n{m['instructor']}" if m.get("instructor") else ""
                    dc.value = f"{m['subject']}{inst_line}\n{m['room']}\n[{m['room_type']}]"
                    dc.fill  = fill(LAB_BG if is_lab else LEC_BG)
                    dc.font  = fnt(bold=True, sz=10,
                                   color=LAB_FG if is_lab else LEC_FG)
                else:
                    dc.fill = fill(WHITE)
                    dc.font = fnt(sz=10)
            row += 1

        # Legend row
        ws.row_dimensions[row].height = 20
        lbl = ws.cell(row=row, column=TC, value="Legend:")
        lbl.font         = fnt(bold=True, sz=10, color=STEEL)
        lbl.fill         = fill(WHITE)
        lbl.alignment    = aln(h="left", wrap=False)
        lbl.border       = box("thin", WHITE)
        lbl.number_format = "@"

        lc = ws.cell(row=row, column=DC["Monday"], value="  Laboratory")
        lc.font         = fnt(bold=True, sz=10, color=LAB_FG)
        lc.fill         = fill(LAB_BG)
        lc.alignment    = aln(h="center", wrap=False)
        lc.border       = box("thin", "93C5FD")
        lc.number_format = "@"

        lc2 = ws.cell(row=row, column=DC["Tuesday"], value="  Lecture Room")
        lc2.font         = fnt(bold=True, sz=10, color=LEC_FG)
        lc2.fill         = fill(LEC_BG)
        lc2.alignment    = aln(h="center", wrap=False)
        lc2.border       = box("thin", "86EFAC")
        lc2.number_format = "@"

        for d in DAYS[2:]:
            cl = ws.cell(row=row, column=DC[d])
            cl.fill   = fill(WHITE)
            cl.border = box("thin", WHITE)
        row += 1

        blank_row(ws, row, TC, T, SEP_BG, 14); row += 1
        blank_row(ws, row, TC, T, WHITE,  10); row += 1

    # ── Print area & freeze panes ─────────────────────────────────
    last_col_letter = get_column_letter(T)
    ws.print_area   = f"A1:{last_col_letter}{row - 1}"
    ws.freeze_panes = ws.cell(row=header_start_row, column=1)

    if "Sheet" in wb.sheetnames and ws.title != "Sheet":
        del wb["Sheet"]

    out = os.path.join(BASE_DIR, "student_schedule_output.xlsx")
    wb.save(out)
    print(f"Excel saved: {out}")
    return out


if __name__ == "__main__":
    print("SmartSched - Generating Student Excel (OnlyOffice compatible)...")
    data = get_data()
    if not data:
        print("No student schedules found in database.")
    else:
        secs = len(set(r["section"] for r in data))
        print(f"Found {len(data)} block(s) from {secs} section(s).")
        generate_excel(data)
        print("Done!")